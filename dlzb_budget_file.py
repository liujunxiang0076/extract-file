import pandas as pd
import openpyxl
import xlrd
import os
import re
import time
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 全局变量用于统计
stats = {
    "total_files": 0,
    "processed_files": 0,
    "matched_budgets": 0,
    "unmatched_budgets": 0,
    "extracted_from_filename": 0,
    "missing_data": {
        "事业部预算编号": 0,
        "合同号": 0,
        "部门（显示值）": 0,
        "单据编号": 0,
        "备注": 0,
        "制单日期": 0,
        "制单人": 0
    }
}

def extract_filenames_to_excel(folder_path, output_file="文件名列表.xlsx", extract_content=False, progress_callback=None):
    """
    提取指定文件夹中所有文件名并保存到Excel文件
    
    Args:
        folder_path: 文件夹路径（字符串或Path对象）
        output_file: 输出Excel文件名
        extract_content: 是否提取Excel文件内容
        progress_callback: 进度回调函数，用于更新GUI进度
    """
    try:
        start_time = time.time()
        
        # 重置统计数据
        global stats
        stats = {
            "total_files": 0,
            "processed_files": 0,
            "matched_budgets": 0,
            "unmatched_budgets": 0,
            "extracted_from_filename": 0,
            "missing_data": {
                "事业部预算编号": 0,
                "合同号": 0,
                "部门（显示值）": 0,
                "单据编号": 0,
                "备注": 0,
                "制单日期": 0,
                "制单人": 0
            }
        }
        
        # 确保folder_path是Path对象
        folder = Path(folder_path)
        
        # 检查文件夹是否存在
        if not folder.exists():
            print(f"错误：文件夹 {folder} 不存在！")
            return
        if not folder.is_dir():
            print(f"错误：{folder} 不是一个文件夹！")
            return
        
        print("开始提取文件名...")
        print(f"目标文件夹：{folder}")
        print("-" * 50)
        
        # 获取所有文件的详细信息
        file_info = []
        excel_files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in ['.xls', '.xlsx']]
        stats["total_files"] = len(excel_files)
        
        # 进度显示
        total_files = len(excel_files)
        processed = 0
        
        for file in excel_files:
            # 更新进度
            processed += 1
            progress_percent = processed / total_files * 100
            
            if progress_callback:
                progress_callback(progress_percent)
                
            if processed % 10 == 0 or processed == total_files:
                print(f"处理进度: {processed}/{total_files} ({progress_percent:.1f}%)")
            
            try:
                # 获取文件统计信息
                stat = file.stat()
                
                # 保存文件的完整路径，用于之后创建超链接
                file_path = str(file.absolute())
                
                file_data = {
                    '文件名': file.stem,
                    '文件路径': file_path,  # 添加文件路径字段
                }
                
                # 如果需要提取文件内容
                if extract_content:
                    try:
                        # 尝试读取Excel文件内容
                        content_data = extract_excel_content(file)
                        # 合并字典
                        file_data.update(content_data)
                    except Exception as e:
                        print(f"警告：无法从文件 {file.name} 提取内容: {e}")
                        # 创建空数据
                        file_data.update({
                            '事业部预算编号': '',
                            '合同号': '',
                            '部门（显示值）': '',
                            '单据编号': '',
                            '备注': '',
                            '制单日期': '',
                            '制单人': ''
                        })
                
                file_info.append(file_data)
                stats["processed_files"] += 1
                
            except Exception as e:
                print(f"处理文件 {file.name} 时出错: {e}")
        
        # 创建DataFrame
        df = pd.DataFrame(file_info)
        
        # 确保列的顺序一致
        column_order = ['文件名']
        if extract_content:
            column_order.extend(['事业部预算编号', '合同号', '部门（显示值）', '单据编号', '备注', '制单日期', '制单人'])
        
        # 添加操作列
        column_order.append('操作')
        
        # 重新排列列
        for col in column_order:
            if col not in df.columns:
                df[col] = ''  # 如果某列不存在，添加空列
        
        # 按指定顺序重排列（但不包括文件路径列，它只用于创建超链接）
        visible_columns = [col for col in column_order if col != '文件路径']
        df_visible = df[visible_columns]
        
        # 确保输出路径在当前项目文件夹中
        current_dir = Path(__file__).parent
        output_path = current_dir / output_file
        
        # 确保输出文件有正确的后缀
        if not output_path.suffix.lower() == '.xlsx':
            output_path = output_path.with_suffix('.xlsx')
        
        # 保存到Excel（不带格式）
        df_visible.to_excel(output_path, index=False, engine='openpyxl')
        
        # 打开工作簿进行格式优化
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.hyperlink import Hyperlink
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hyperlink_font = Font(name='微软雅黑', size=10, color="0563C1", underline="single")
        
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头样式
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 获取操作列的列号
        operation_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == '操作':
                operation_col = col
                break
        
        # 设置数据区域样式和添加超链接
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
                
                # 为操作列添加超链接和文本
                if col == operation_col:
                    # 获取当前行对应的文件路径
                    file_path = df.iloc[row-2]['文件路径']
                    
                    # 设置操作列的文本和样式
                    cell.value = "打开文件"
                    cell.font = hyperlink_font
                    
                    # 添加超链接
                    cell.hyperlink = file_path
        
        # 自动调整列宽
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            # 获取该列最长内容的长度
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            # 设置列宽（根据内容长度计算，中文字符宽度需要调整）
            adjusted_width = max_length * 1.2 + 4  # 中文字符宽度调整系数
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 添加统计信息到新工作表
        ws_stats = wb.create_sheet(title="统计信息")
        
        # 添加标题
        ws_stats['A1'] = "提取统计信息"
        ws_stats.merge_cells('A1:B1')
        ws_stats['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        ws_stats['A1'].alignment = Alignment(horizontal='center')
        
        # 添加统计数据
        ws_stats['A3'] = "总文件数"
        ws_stats['B3'] = stats["total_files"]
        ws_stats['A4'] = "成功处理文件数"
        ws_stats['B4'] = stats["processed_files"]
        ws_stats['A5'] = "预算编号匹配文件数"
        ws_stats['B5'] = stats["matched_budgets"]
        ws_stats['A6'] = "预算编号不匹配文件数"
        ws_stats['B6'] = stats["unmatched_budgets"]
        ws_stats['A7'] = "从文件名提取预算编号数"
        ws_stats['B7'] = stats["extracted_from_filename"]
        
        # 缺失数据统计
        ws_stats['A9'] = "缺失数据统计"
        ws_stats.merge_cells('A9:B9')
        ws_stats['A9'].font = Font(bold=True)
        
        row = 10
        for field, count in stats["missing_data"].items():
            ws_stats[f'A{row}'] = f"缺失{field}的文件数"
            ws_stats[f'B{row}'] = count
            row += 1
        
        # 设置统计表格的列宽
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
        
        # 保存格式化后的Excel
        wb.save(output_path)
        
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        print(f"完成：共提取 {len(file_info)} 个文件的详细信息")
        print(f"处理时间：{elapsed_time:.2f}秒")
        print(f"已保存到：{output_path.absolute()}")
        print("-" * 50)
        print("统计信息:")
        print(f"  总文件数: {stats['total_files']}")
        print(f"  成功处理文件数: {stats['processed_files']}")
        print(f"  预算编号匹配文件数: {stats['matched_budgets']}")
        print(f"  预算编号不匹配文件数: {stats['unmatched_budgets']}")
        print(f"  从文件名提取预算编号数: {stats['extracted_from_filename']}")
        print("  缺失数据统计:")
        for field, count in stats["missing_data"].items():
            print(f"    缺失{field}的文件数: {count}")
        print("-" * 50)
        
        return file_info
            
    except Exception as e:
        print(f"出错：{e}")
        import traceback
        traceback.print_exc()
        return []

def normalize_budget_id(text):
    """
    标准化预算编号格式，去除多余空格和特殊字符
    """
    if not text:
        return ""
        
    # 去除多余空格和非法字符
    text = re.sub(r'\s+', '', text)
    
    # 尝试匹配标准格式 WZ-FJ-YYYYMM-NNN
    pattern = r'([A-Z]{1,2})[-_]?([A-Z]{1,2})[-_]?(\d{6})[-_]?(\d{3})'
    match = re.search(pattern, text)
    if match:
        # 重新格式化为标准格式
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}-{match.group(4)}"
    
    return text

def extract_excel_content(file_path):
    """
    从Excel文件中提取特定内容
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        包含提取内容的字典
    """
    try:
        file_ext = Path(file_path).suffix.lower()
        # 调整字典顺序，确保与提取顺序一致
        result = {
            '事业部预算编号': '',
            '合同号': '',
            '部门（显示值）': '',
            '单据编号': '',
            '备注': '',
            '制单日期': '',
            '制单人': ''
        }
        
        if file_ext == '.xlsx':
            # 使用openpyxl读取.xlsx文件
            return extract_with_openpyxl(file_path, result)
        elif file_ext == '.xls':
            # 使用xlrd读取.xls文件
            return extract_with_xlrd(file_path, result)
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}")
    
    except Exception as e:
        print(f"提取文件 {file_path} 内容时出错: {e}")
        return {
            '事业部预算编号': '',
            '合同号': '',
            '部门（显示值）': '',
            '单据编号': '',
            '备注': '',
            '制单日期': '',
            '制单人': ''
        }

def find_value_by_keyword(worksheet, keywords, max_rows=100, max_cols=30):
    """在工作表中查找关键字并返回相应的值"""
    # 标准化关键字
    normalized_keywords = [k.strip().lower() for k in keywords if k]
    
    # 识别表格结构的函数
    def detect_table_structure(ws):
        """检测工作表中的表格结构，返回可能的表头行和表头列"""
        headers_row = None
        headers_col = None
        header_candidates = ['序号', '项目', '名称', '编号', '型号', '代码']
        
        # 对openpyxl工作表
        if hasattr(ws, 'iter_rows'):
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=10), 1):
                header_count = 0
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.strip().lower()
                        if any(h.lower() in cell_text for h in header_candidates):
                            header_count += 1
                if header_count >= 3:  # 如果一行中有3个以上可能的表头
                    headers_row = row_idx
                    break
        # 对xlrd工作表
        else:
            for row_idx in range(min(30, ws.nrows)):
                header_count = 0
                for col_idx in range(min(10, ws.ncols)):
                    cell_value = ws.cell_value(row_idx, col_idx)
                    if cell_value and isinstance(cell_value, str):
                        cell_text = cell_value.strip().lower()
                        if any(h.lower() in cell_text for h in header_candidates):
                            header_count += 1
                if header_count >= 3:
                    headers_row = row_idx
                    break
        
        return headers_row
    
    # 尝试检测表格结构
    table_header_row = detect_table_structure(worksheet)
    
    # 对openpyxl工作表的处理
    if hasattr(worksheet, 'iter_rows'):
        # 1. 首先搜索标题区域 (通常位于前20行)
        # 先查找单据编号、合同号等关键信息，这些通常位于文件上方
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, max_rows), max_col=max_cols), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip().lower()
                    # 寻找精确匹配的单元格
                    for keyword in normalized_keywords:
                        # 精确匹配关键字
                        if (cell_text == keyword or 
                            cell_text == keyword + "：" or 
                            cell_text == keyword + ":" or
                            cell_text.endswith(keyword)):
                            
                            # 优先检查右侧单元格
                            if col_idx < max_cols:
                                right_cell = worksheet.cell(row=row_idx, column=col_idx+1)
                                if right_cell.value:
                                    return str(right_cell.value).strip()
                            
                            # 其次检查下方单元格
                            if row_idx < max_rows:
                                below_cell = worksheet.cell(row=row_idx+1, column=col_idx)
                                if below_cell.value:
                                    return str(below_cell.value).strip()
                                    
                            # 再检查右下方单元格
                            if col_idx < max_cols and row_idx < max_rows:
                                diag_cell = worksheet.cell(row=row_idx+1, column=col_idx+1)
                                if diag_cell.value:
                                    return str(diag_cell.value).strip()
        
        # 2. 如果检测到表格结构，尝试在表格中搜索关键字
        if table_header_row:
            # 获取表头
            headers = []
            for col in range(1, min(max_cols, worksheet.max_column) + 1):
                header_cell = worksheet.cell(row=table_header_row, column=col)
                if header_cell.value:
                    headers.append((col, str(header_cell.value).strip().lower()))
            
            # 查找匹配关键字的列
            target_cols = []
            for col_idx, header in headers:
                for keyword in normalized_keywords:
                    if keyword in header:
                        target_cols.append(col_idx)
                        break
            
            # 如果找到匹配的列，返回该列中第一个非空单元格的值
            if target_cols:
                for row in range(table_header_row + 1, min(max_rows, worksheet.max_row) + 1):
                    for col in target_cols:
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value:
                            return str(cell.value).strip()
                            
        # 3. 全文搜索包含关键字的单元格
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip().lower()
                    for keyword in normalized_keywords:
                        if keyword in cell_text:
                            # 查找同一行中其他单元格是否包含值
                            for check_col in range(1, min(max_cols, worksheet.max_column) + 1):
                                if check_col != col_idx:
                                    check_cell = worksheet.cell(row=row_idx, column=check_col)
                                    if check_cell.value and isinstance(check_cell.value, str) and len(check_cell.value.strip()) > 0:
                                        # 过滤掉可能的表头或标签
                                        check_text = check_cell.value.strip().lower()
                                        if not any(k.lower() in check_text for k in normalized_keywords):
                                            return str(check_cell.value).strip()
                            
                            # 尝试检查右侧单元格
                            if col_idx < max_cols:
                                right_cell = worksheet.cell(row=row_idx, column=col_idx+1)
                                if right_cell.value:
                                    return str(right_cell.value).strip()
                            
                            # 检查下方单元格
                            if row_idx < max_rows:
                                below_cell = worksheet.cell(row=row_idx+1, column=col_idx)
                                if below_cell.value:
                                    return str(below_cell.value).strip()
                                    
                            # 检查合并单元格
                            try:
                                for merged_range in worksheet.merged_cells.ranges:
                                    min_col, min_row, max_col, max_row = merged_range.bounds
                                    if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
                                        # 如果当前单元格在一个合并区域内
                                        next_col = max_col + 1
                                        if next_col <= worksheet.max_column:
                                            right_cell = worksheet.cell(row=row_idx, column=next_col)
                                            if right_cell.value:
                                                return str(right_cell.value).strip()
                            except:
                                pass
        
        # 4. 最后搜索特定模式，如"单据编号WZBD20240425"
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).strip()
                    # 查找"关键字+值"的模式
                    for keyword in keywords:
                        keyword_lower = keyword.lower()
                        if keyword_lower in cell_text.lower():
                            # 提取关键字后面的内容
                            parts = re.split(f"{re.escape(keyword_lower)}[：:]*", cell_text.lower(), flags=re.IGNORECASE)
                            if len(parts) > 1 and parts[1].strip():
                                # 提取实际的值（考虑大小写）
                                start_idx = cell_text.lower().find(keyword_lower) + len(keyword_lower)
                                # 跳过可能的冒号和空格
                                while start_idx < len(cell_text) and (cell_text[start_idx] in [':', '：', ' ']):
                                    start_idx += 1
                                return cell_text[start_idx:].strip()
    
    # 对xlrd工作表的处理
    else:
        # 1. 首先搜索标题区域 (通常位于前20行)
        for row_idx in range(min(20, min(max_rows, worksheet.nrows))):
            for col_idx in range(min(max_cols, worksheet.ncols)):
                cell_value = worksheet.cell_value(row_idx, col_idx)
                if cell_value:
                    cell_text = str(cell_value).strip().lower()
                    for keyword in normalized_keywords:
                        # 精确匹配关键字
                        if (cell_text == keyword or 
                            cell_text == keyword + "：" or 
                            cell_text == keyword + ":" or
                            cell_text.endswith(keyword)):
                            
                            # 优先检查右侧单元格
                            if col_idx + 1 < worksheet.ncols:
                                right_value = worksheet.cell_value(row_idx, col_idx + 1)
                                if right_value:
                                    return str(right_value).strip()
                            
                            # 其次检查下方单元格
                            if row_idx + 1 < worksheet.nrows:
                                below_value = worksheet.cell_value(row_idx + 1, col_idx)
                                if below_value:
                                    return str(below_value).strip()
                                    
                            # 再检查右下方单元格
                            if col_idx + 1 < worksheet.ncols and row_idx + 1 < worksheet.nrows:
                                diag_value = worksheet.cell_value(row_idx + 1, col_idx + 1)
                                if diag_value:
                                    return str(diag_value).strip()
        
        # 2. 如果检测到表格结构，尝试在表格中搜索关键字
        if table_header_row:
            # 获取表头
            headers = []
            for col_idx in range(min(max_cols, worksheet.ncols)):
                header_value = worksheet.cell_value(table_header_row, col_idx)
                if header_value:
                    headers.append((col_idx, str(header_value).strip().lower()))
            
            # 查找匹配关键字的列
            target_cols = []
            for col_idx, header in headers:
                for keyword in normalized_keywords:
                    if keyword in header:
                        target_cols.append(col_idx)
                        break
            
            # 如果找到匹配的列，返回该列中第一个非空单元格的值
            if target_cols:
                for row_idx in range(table_header_row + 1, min(max_rows, worksheet.nrows)):
                    for col_idx in target_cols:
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            return str(cell_value).strip()
        
        # 3. 全文搜索包含关键字的单元格
        for row_idx in range(min(max_rows, worksheet.nrows)):
            for col_idx in range(min(max_cols, worksheet.ncols)):
                cell_value = worksheet.cell_value(row_idx, col_idx)
                if cell_value:
                    cell_text = str(cell_value).strip().lower()
                    for keyword in normalized_keywords:
                        if keyword in cell_text:
                            # 查找同一行中其他单元格是否包含值
                            for check_col in range(min(max_cols, worksheet.ncols)):
                                if check_col != col_idx:
                                    check_value = worksheet.cell_value(row_idx, check_col)
                                    if check_value and isinstance(check_value, str) and len(str(check_value).strip()) > 0:
                                        # 过滤掉可能的表头或标签
                                        check_text = str(check_value).strip().lower()
                                        if not any(k.lower() in check_text for k in normalized_keywords):
                                            return str(check_value).strip()
                            
                            # 检查右侧单元格
                            if col_idx + 1 < worksheet.ncols:
                                right_value = worksheet.cell_value(row_idx, col_idx + 1)
                                if right_value:
                                    return str(right_value).strip()
                            
                            # 检查下方单元格
                            if row_idx + 1 < worksheet.nrows:
                                below_value = worksheet.cell_value(row_idx + 1, col_idx)
                                if below_value:
                                    return str(below_value).strip()
        
        # 4. 最后搜索特定模式，如"单据编号WZBD20240425"
        for row_idx in range(min(max_rows, worksheet.nrows)):
            for col_idx in range(min(max_cols, worksheet.ncols)):
                cell_value = worksheet.cell_value(row_idx, col_idx)
                if cell_value and isinstance(cell_value, str):
                    cell_text = str(cell_value).strip()
                    # 查找"关键字+值"的模式
                    for keyword in keywords:
                        keyword_lower = keyword.lower()
                        if keyword_lower in cell_text.lower():
                            # 提取关键字后面的内容
                            parts = re.split(f"{re.escape(keyword_lower)}[：:]*", cell_text.lower(), flags=re.IGNORECASE)
                            if len(parts) > 1 and parts[1].strip():
                                # 提取实际的值（考虑大小写）
                                start_idx = cell_text.lower().find(keyword_lower) + len(keyword_lower)
                                # 跳过可能的冒号和空格
                                while start_idx < len(cell_text) and (cell_text[start_idx] in [':', '：', ' ']):
                                    start_idx += 1
                                return cell_text[start_idx:].strip()
    
    return ""

def find_value_by_coordinate(worksheet, col, row):
    """
    根据坐标获取单元格的值
    
    Args:
        worksheet: 工作表对象
        col: 列坐标 (A, B, C, ...)
        row: 行坐标 (1, 2, 3, ...)
        
    Returns:
        单元格的值或空字符串
    """
    try:
        # 转换列坐标为数字 (A->1, B->2, ...)
        col_idx = ord(col.upper()) - ord('A') + 1
        
        # 对openpyxl工作表
        if hasattr(worksheet, 'cell'):
            cell = worksheet.cell(row=row, column=col_idx)
            if cell.value:
                return str(cell.value).strip()
        # 对xlrd工作表
        else:
            row_idx = row - 1  # xlrd行索引从0开始
            col_idx = col_idx - 1  # xlrd列索引从0开始
            if row_idx < worksheet.nrows and col_idx < worksheet.ncols:
                value = worksheet.cell_value(row_idx, col_idx)
                if value:
                    return str(value).strip()
    except Exception as e:
        print(f"根据坐标获取单元格值时出错: {e}")
    
    return ""

def find_value_in_column(worksheet, col_letter, keyword, max_rows=100):
    """
    在指定列中查找包含关键字的行，并返回该行对应列的值
    
    Args:
        worksheet: 工作表对象
        col_letter: 列字母 (如 'A', 'D')
        keyword: 要查找的关键字
        max_rows: 最大搜索行数
        
    Returns:
        找到的值或空字符串
    """
    try:
        col_idx = ord(col_letter.upper()) - ord('A') + 1
        keyword_lower = keyword.lower()
        
        # 对openpyxl工作表
        if hasattr(worksheet, 'iter_rows'):
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_rows), 1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and keyword_lower in cell.value.lower():
                        # 找到关键字，返回同一行指定列的值
                        target_cell = worksheet.cell(row=row_idx, column=col_idx)
                        if target_cell.value:
                            return str(target_cell.value).strip()
        # 对xlrd工作表
        else:
            col_idx = col_idx - 1  # xlrd列索引从0开始
            for row_idx in range(min(max_rows, worksheet.nrows)):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row_idx, col)
                    if cell_value and isinstance(cell_value, str) and keyword_lower in cell_value.lower():
                        # 找到关键字，返回同一行指定列的值
                        if col_idx < worksheet.ncols:
                            value = worksheet.cell_value(row_idx, col_idx)
                            if value:
                                return str(value).strip()
    except Exception as e:
        print(f"在列中查找值时出错: {e}")
    
    return ""

def clean_extracted_value(value, field_name):
    """
    清理提取的值，移除列标题和多余空格
    
    Args:
        value: 提取的原始值
        field_name: 字段名称
    
    Returns:
        清理后的值
    """
    if not value:
        return ""
    
    # 去除多余空格
    value = value.strip()
    
    # 检查是否包含字段名称，并去除
    field_patterns = [
        f"{field_name}[：:]*\\s*",
        f"{field_name.replace('（', '(').replace('）', ')')}[：:]*\\s*"
    ]
    
    for pattern in field_patterns:
        value = re.sub(pattern, "", value, flags=re.IGNORECASE)
    
    return value.strip()

def extract_with_openpyxl(file_path, result):
    """使用openpyxl提取.xlsx文件内容"""
    try:
        # 使用openpyxl读取Excel文件
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 1. 根据坐标查找固定位置的值（根据截图中的位置）
        # A列下的"事业部预算编号"数据实际上是G列的合同号数据
        result['合同号'] = find_value_by_coordinate(ws, 'A', 4)
        
        # G列下的"合同号"数据实际上是A列的事业部预算编号数据
        result['事业部预算编号'] = find_value_by_coordinate(ws, 'G', 4)
        
        # A列第5行是部门信息
        result['部门（显示值）'] = find_value_by_coordinate(ws, 'A', 5)
        
        # A列第6行是单据编号（这个没问题）
        result['单据编号'] = find_value_by_coordinate(ws, 'A', 6)
        
        # G列第6行是备注信息（这个没问题）
        result['备注'] = find_value_by_coordinate(ws, 'G', 6)
        
        # 查找制单日期和制单人信息
        result['制单日期'] = find_value_in_column(ws, 'G', '制单日期')
        result['制单人'] = find_value_in_column(ws, 'H', '制单人')
        
        # 2. 如果以上方法未能提取到全部信息，尝试使用关键字搜索
        if not result['事业部预算编号']:
            result['事业部预算编号'] = find_value_by_keyword(ws, ['事业部预算编号'])
        
        if not result['合同号']:
            result['合同号'] = find_value_by_keyword(ws, ['合同号'])
        
        if not result['部门（显示值）']:
            result['部门（显示值）'] = find_value_by_keyword(ws, ['部门（显示值）', '部门(显示值)', '部门', '使用部门', '申请部门', '所属部门', '责任部门'])
        
        if not result['单据编号']:
            result['单据编号'] = find_value_by_keyword(ws, ['单据编号', '单据号', '凭证号', '凭证编号', '发票号', '发票编号', '申请单号'])
        
        if not result['备注']:
            result['备注'] = find_value_by_keyword(ws, ['备注', '备注说明', '说明', '项目说明', '其他说明', '补充说明', '附注'])
        
        if not result['制单日期']:
            result['制单日期'] = find_value_by_keyword(ws, ['制单日期'])
        
        if not result['制单人']:
            result['制单人'] = find_value_by_keyword(ws, ['制单人'])
        
        # 3. 如果常规方法未能提取到全部信息，尝试扫描整个表格寻找特定模式
        if not all(result.values()):
            # 尝试扫描整个表格，寻找包含关键信息的行
            budget_pattern = r'WZ[-_]?FJ[-_]?(\d{6})[-_]?(\d{3})'
            document_pattern = r'WZBD(\d{8})'
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, max_col=20), 1):
                row_text = ' '.join(str(cell.value or '').strip() for cell in row)
                
                # 查找并提取预算编号
                if not result['事业部预算编号']:
                    budget_match = re.search(budget_pattern, row_text)
                    if budget_match:
                        # 提取完整匹配内容
                        full_match = budget_match.group(0)
                        # 规范化格式
                        result['事业部预算编号'] = normalize_budget_id(full_match)
                
                # 查找并提取单据编号
                if not result['单据编号']:
                    doc_match = re.search(document_pattern, row_text)
                    if doc_match:
                        # 提取完整匹配内容
                        full_match = doc_match.group(0)
                        result['单据编号'] = full_match
                        
                # 尝试识别部门信息
                if not result['部门（显示值）'] and ('辅机' in row_text and '事业部' in row_text):
                    result['部门（显示值）'] = '辅机事业部'
                    
                # 查找合同信息（通常比较长，可能包含"锅炉"、"滤网"等关键词）
                if not result['合同号'] and ('锅炉' in row_text or '滤网' in row_text or '采购' in row_text):
                    # 提取可能的合同描述
                    # 排除掉明显的编号和固定文本
                    clean_text = re.sub(budget_pattern, '', row_text)
                    clean_text = re.sub(document_pattern, '', clean_text)
                    clean_text = re.sub(r'部门[（(]显示值[)）]', '', clean_text)
                    clean_text = re.sub(r'事业部预算编号|预算编号|单据编号|部门|备注', '', clean_text)
                    
                    # 清理多余空格
                    clean_text = re.sub(r'\s+', ' ', clean_text).strip()
                    
                    if clean_text and len(clean_text) > 3:  # 至少有一些有意义的文本
                        result['合同号'] = clean_text
        
        # 清理提取的数据
        for key in result:
            if result[key]:
                result[key] = clean_extracted_value(result[key], key)
        
        # 标准化最终结果
        if result['事业部预算编号']:
            result['事业部预算编号'] = normalize_budget_id(result['事业部预算编号'])
            
        # 清理单据编号中可能的前缀
        if result['单据编号']:
            # 提取WZBD后面的数字部分
            doc_match = re.search(r'WZBD(\d+)', result['单据编号'])
            if doc_match:
                result['单据编号'] = f"WZBD{doc_match.group(1)}"
        
        # 记录缺失数据统计
        for field, value in result.items():
            if not value:
                stats["missing_data"][field] += 1
        
        # 如果没有找到事业部预算编号，尝试从文件名提取
        if not result['事业部预算编号']:
            file_stem = Path(file_path).stem
            # 尝试从文件名中提取预算编号格式
            budget_id_match = re.search(r'([A-Z]{1,2})[-_]?([A-Z]{1,2})[-_]?(\d{6})[-_]?(\d{3})', file_stem)
            if budget_id_match:
                extracted_id = f"{budget_id_match.group(1)}-{budget_id_match.group(2)}-{budget_id_match.group(3)}-{budget_id_match.group(4)}"
                result['事业部预算编号'] = extracted_id
                stats["extracted_from_filename"] += 1
                print(f"✓ 从文件名 {file_stem} 提取预算编号: {extracted_id}")
        
        # 验证事业部预算编号与文件名的关系（更宽松的匹配）
        file_stem = Path(file_path).stem
        if result['事业部预算编号']:
            # 标准化文件名中的预算编号格式
            normalized_stem = normalize_budget_id(file_stem)
            
            # 提取纯数字部分进行比较
            file_numbers = re.sub(r'[^0-9]', '', normalized_stem)
            budget_numbers = re.sub(r'[^0-9]', '', result['事业部预算编号'])
            
            # 如果数字部分包含关系，也算匹配
            if (budget_numbers in file_numbers) or (file_numbers in budget_numbers):
                print(f"√ 文件 {file_stem} 的事业部预算编号验证通过")
                stats["matched_budgets"] += 1
            else:
                print(f"! 警告：文件 {file_stem} 的事业部预算编号与文件名不匹配")
                stats["unmatched_budgets"] += 1
                # 使用文件名作为预算编号
                if not result['事业部预算编号'] and normalized_stem:
                    result['事业部预算编号'] = normalized_stem
                    print(f"  > 已使用文件名 {normalized_stem} 作为预算编号")
        
        return result
    
    except Exception as e:
        print(f"使用openpyxl提取 {file_path} 时出错: {e}")
        for field in result:
            stats["missing_data"][field] += 1
        return result

def extract_with_xlrd(file_path, result):
    """使用xlrd提取.xls文件内容"""
    try:
        # 使用xlrd读取Excel文件
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)  # 获取第一个工作表
        
        # 1. 根据坐标查找固定位置的值（根据截图中的位置）
        # A列下的"事业部预算编号"数据实际上是G列的合同号数据
        result['合同号'] = find_value_by_coordinate(ws, 'A', 4)
        
        # G列下的"合同号"数据实际上是A列的事业部预算编号数据
        result['事业部预算编号'] = find_value_by_coordinate(ws, 'G', 4)
        
        # A列第5行是部门信息
        result['部门（显示值）'] = find_value_by_coordinate(ws, 'A', 5)
        
        # A列第6行是单据编号（这个没问题）
        result['单据编号'] = find_value_by_coordinate(ws, 'A', 6)
        
        # G列第6行是备注信息（这个没问题）
        result['备注'] = find_value_by_coordinate(ws, 'G', 6)
        
        # 查找制单日期和制单人信息
        result['制单日期'] = find_value_in_column(ws, 'G', '制单日期')
        result['制单人'] = find_value_in_column(ws, 'H', '制单人')
        
        # 2. 如果以上方法未能提取到全部信息，尝试使用关键字搜索
        if not result['事业部预算编号']:
            result['事业部预算编号'] = find_value_by_keyword(ws, ['事业部预算编号'])
        
        if not result['合同号']:
            result['合同号'] = find_value_by_keyword(ws, ['合同号'])
        
        if not result['部门（显示值）']:
            result['部门（显示值）'] = find_value_by_keyword(ws, ['部门（显示值）', '部门(显示值)', '部门', '使用部门', '申请部门', '所属部门', '责任部门'])
        
        if not result['单据编号']:
            result['单据编号'] = find_value_by_keyword(ws, ['单据编号', '单据号', '凭证号', '凭证编号', '发票号', '发票编号', '申请单号'])
        
        if not result['备注']:
            result['备注'] = find_value_by_keyword(ws, ['备注', '备注说明', '说明', '项目说明', '其他说明', '补充说明', '附注'])
        
        if not result['制单日期']:
            result['制单日期'] = find_value_by_keyword(ws, ['制单日期'])
        
        if not result['制单人']:
            result['制单人'] = find_value_by_keyword(ws, ['制单人'])
        
        # 清理提取的数据
        for key in result:
            if result[key]:
                result[key] = clean_extracted_value(result[key], key)
        
        # ... 其余代码保持不变
        
        return result
    
    except xlrd.biffh.XLRDError as e:
        # ... 现有代码保持不变
        return result
    
    except Exception as e:
        print(f"使用xlrd提取 {file_path} 时出错: {e}")
        for field in result:
            stats["missing_data"][field] += 1
        return result

def create_test_files():
    """创建测试文件"""
    # 创建测试目录
    test_dir = Path(__file__).parent / "test_run"
    test_dir.mkdir(exist_ok=True)
    
    # 创建第一个.xlsx文件 - 简单结构
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 添加标题行
    ws['A1'] = "预算单"
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 添加内容 - 按照实际位置设置数据
    # A4 存放合同号数据
    ws['A4'] = "合同号：太平洋1#锅炉形滤网"
    # G4 存放事业部预算编号数据
    ws['G4'] = "事业部预算编号：WZ-FJ-202406-032"
    # A5 存放部门信息
    ws['A5'] = "部门（显示值）：辅机事业部"
    # A6 存放单据编号
    ws['A6'] = "单据编号：WZBD20240197"
    # G6 存放备注
    ws['G6'] = "备注：王志中"
    # 制单信息（通常在G列和H列）
    ws['G10'] = "制单日期：2024-06-10"
    ws['H10'] = "制单人：李四"
    
    # 保存文件
    test_file_path = test_dir / "WZ-FJ-202406-032.xlsx"
    wb.save(test_file_path)
    
    # 创建第二个.xlsx文件（不同格式）
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    
    # 添加表格形式的数据
    ws2['A1'] = "项目信息表"
    ws2.merge_cells('A1:F1')
    
    # 添加表头 - 确保顺序与导出一致
    headers = ["序号", "项目编号", "项目内容", "申请部门", "单据号", "说明"]
    for i, header in enumerate(headers, 1):
        ws2.cell(row=3, column=i).value = header
    
    # 添加一行数据 - 确保顺序与导出一致
    row_data = ["1", "WZ-FJ-201412-039", "设备采购", "辅机事业部", "WZBD2024042", "测试备注"]
    for i, data in enumerate(row_data, 1):
        ws2.cell(row=4, column=i).value = data
    
    # 添加制单信息
    ws2['G15'] = "制单日期：2024-05-15"
    ws2['H15'] = "制单人：张三"
    
    # 保存第二个文件
    test_file_path2 = test_dir / "WZ-FJ-201412-039.xlsx"
    wb2.save(test_file_path2)
    
    # 创建第三个.xlsx文件 - 模拟截图中的表格格式
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    
    # 设置样式
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 第一部分 - 模拟截图中的表格头
    # 添加表头（按照实际顺序）
    headers = ["文件名", "合同号", "事业部预算编号", "部门（显示值）", "单据编号", "备注", "制单日期", "制单人"]
    for i, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=i)
        cell.value = header
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加几行数据（按照正确顺序）
    data_rows = [
        ["WZ-FJ-201412-039", "三峡项目3、4号机组LOTA装置", "WZ-FJ-201412-039", "辅机事业部", "WZBD20240425", "滤网", "2024-04-15", "王五"],
        ["WZ-FJ-202304-023", "康江1#2#锅炉水系统平板滤网", "WZ-FJ-202304-023", "辅机事业部", "WZBD20240073", "稳华", "2024-03-20", "赵六"],
        ["WZ-FJ-202309-007", "太平岭1#2级滤网", "WZ-FJ-202309-007", "辅机事业部", "WZBD20240056", "王志中", "2024-02-28", "李四"]
    ]
    
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # 保存第三个文件 - 模拟表格式样
    complex_file_path = test_dir / "模拟复杂表格.xlsx"
    wb3.save(complex_file_path)
    
    # 创建第四个.xlsx文件 - 模拟第二张截图中的预算单格式
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    
    # 添加标题
    ws4['A2'] = "预算单"
    ws4.merge_cells('A2:H2')
    ws4['A2'].alignment = Alignment(horizontal='center')
    ws4['A2'].font = Font(size=14, bold=True)
    
    # 添加生产性物资采购预算单
    ws4['A3'] = "生产性物资采购预算单"
    ws4.merge_cells('A3:H3')
    ws4['A3'].alignment = Alignment(horizontal='left')
    ws4['A3'].font = Font(size=12, bold=True)
    
    # 添加项目信息
    ws4['A4'] = "合同号：集江1#2#锅炉水系统平板滤网4、海水取水"
    ws4.merge_cells('A4:C4')
    ws4['G4'] = "事业部预算编号：WZ-FJ-202312-033"
    
    ws4['A5'] = "部门(显示值)：辅机事业部"
    ws4.merge_cells('A5:C5')
    
    ws4['A6'] = "单据编号：WZBD20240005"
    ws4.merge_cells('A6:C6')
    ws4['G6'] = "备注：平板滤网"
    
    # 添加制单信息（放在底部）
    ws4['G25'] = "制单日期：2024-05-18"
    ws4['H25'] = "制单人：陈明"
    
    # 添加表格头
    table_headers = ["序号", "存货编码", "存货名称", "规格型号", "材质", "单位", "预算数量", "技术标准"]
    for i, header in enumerate(table_headers, 1):
        cell = ws4.cell(row=8, column=i)
        cell.value = header
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 保存第四个文件
    budget_form_path = test_dir / "WZ-FJ-202312-033_预算单.xlsx"
    wb4.save(budget_form_path)
    
    # 创建.xls格式的测试文件（使用xlwt直接创建）
    try:
        import xlwt
        
        wb5 = xlwt.Workbook()
        ws5 = wb5.add_sheet('Sheet1')
        
        # 添加内容，模拟老式Excel - 按照正确的位置和字段
        ws5.write(3, 0, "合同号：")
        ws5.write(3, 1, "备品备件采购")
        ws5.write(3, 6, "事业部预算编号：")
        ws5.write(3, 7, "WZ-FJ-202306-128")
        ws5.write(4, 0, "部门：")
        ws5.write(4, 1, "维修部")
        ws5.write(5, 0, "单据编号：")
        ws5.write(5, 1, "WZBD20230356")
        ws5.write(5, 6, "备注：")
        ws5.write(5, 7, "年度预算")
        ws5.write(24, 6, "制单日期：")
        ws5.write(24, 7, "2024-03-15")
        ws5.write(24, 8, "制单人：")
        ws5.write(24, 9, "张三")
        
        # 保存为.xls格式
        xls_path = test_dir / "WZ-FJ-202306-128.xls"
        wb5.save(xls_path)
        print(f"已创建.xls测试文件: {xls_path}")
    except ImportError:
        print("警告: 未安装xlwt库，无法创建.xls测试文件")
        print("请安装xlwt: pip install xlwt")
    except Exception as e:
        print(f"创建.xls测试文件时出错: {e}")
    
    print(f"已创建测试文件目录: {test_dir}")
    return test_dir

def create_gui():
    """
    创建简单的GUI界面，方便用户选择文件夹和操作
    """
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
        import threading
        
        # 创建主窗口
        root = tk.Tk()
        root.title("预算文件提取工具")
        root.geometry("600x450")
        root.resizable(True, True)
        
        # 设置样式
        style = ttk.Style()
        style.configure("TButton", font=("微软雅黑", 10))
        style.configure("TLabel", font=("微软雅黑", 10))
        style.configure("TCheckbutton", font=("微软雅黑", 10))
        style.configure("TEntry", font=("微软雅黑", 10))
        
        # 创建框架
        main_frame = ttk.Frame(root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建标题
        title_label = ttk.Label(main_frame, text="Excel预算文件信息提取工具", font=("微软雅黑", 14, "bold"))
        title_label.pack(pady=10)
        
        # 创建输入框架
        input_frame = ttk.LabelFrame(main_frame, text="输入设置", padding=10)
        input_frame.pack(fill=tk.X, pady=10)
        
        # 文件夹选择
        folder_frame = ttk.Frame(input_frame)
        folder_frame.pack(fill=tk.X, pady=5)
        
        folder_label = ttk.Label(folder_frame, text="Excel文件夹:")
        folder_label.pack(side=tk.LEFT, padx=5)
        
        folder_var = tk.StringVar()
        folder_entry = ttk.Entry(folder_frame, textvariable=folder_var, width=40)
        folder_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        def browse_folder():
            folder_path = filedialog.askdirectory(title="选择Excel文件夹")
            if folder_path:
                folder_var.set(folder_path)
        
        browse_button = ttk.Button(folder_frame, text="浏览...", command=browse_folder)
        browse_button.pack(side=tk.LEFT, padx=5)
        
        # 输出文件名
        output_frame = ttk.Frame(input_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        output_label = ttk.Label(output_frame, text="输出文件名:")
        output_label.pack(side=tk.LEFT, padx=5)
        
        output_var = tk.StringVar(value="预算文件列表.xlsx")
        output_entry = ttk.Entry(output_frame, textvariable=output_var, width=40)
        output_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 提取内容选项
        extract_var = tk.BooleanVar(value=True)
        extract_check = ttk.Checkbutton(input_frame, text="提取Excel文件内容", variable=extract_var)
        extract_check.pack(anchor=tk.W, pady=5)
        
        # 操作按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # 进度条
        progress_frame = ttk.LabelFrame(main_frame, text="处理进度", padding=10)
        progress_frame.pack(fill=tk.X, pady=10)
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_frame, variable=progress_var, maximum=100)
        progress_bar.pack(fill=tk.X, pady=5)
        
        progress_label = ttk.Label(progress_frame, text="就绪")
        progress_label.pack(anchor=tk.W)
        
        # 日志框架
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 创建文本框和滚动条
        log_text = tk.Text(log_frame, height=10, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(log_frame, command=log_text.yview)
        log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 自定义输出重定向
        class TextRedirector:
            def __init__(self, text_widget):
                self.text_widget = text_widget
                self.buffer = ""
                
            def write(self, string):
                self.buffer += string
                self.text_widget.insert(tk.END, string)
                self.text_widget.see(tk.END)
                self.text_widget.update_idletasks()
                
            def flush(self):
                pass
        
        # 进度更新函数
        def update_progress(value):
            root.after(0, lambda: progress_var.set(value))
            root.after(0, lambda: progress_label.config(text=f"处理进度: {value:.1f}%"))
        
        # 处理函数
        def process_files():
            folder_path = folder_var.get()
            output_file = output_var.get()
            extract_content = extract_var.get()
            
            if not folder_path:
                messagebox.showerror("错误", "请选择Excel文件夹!")
                return
                
            # 禁用按钮，避免重复点击
            start_button.config(state=tk.DISABLED)
            test_button.config(state=tk.DISABLED)
            
            # 清空日志
            log_text.delete(1.0, tk.END)
            
            # 更新进度信息
            progress_label.config(text="正在处理...")
            progress_var.set(0)
            
            # 重定向标准输出到文本框
            import sys
            original_stdout = sys.stdout
            sys.stdout = TextRedirector(log_text)
            
            def run_extraction():
                try:
                    # 运行提取函数，传入进度回调
                    extract_filenames_to_excel(folder_path, output_file, extract_content, progress_callback=update_progress)
                    
                    # 完成后在主线程更新UI
                    root.after(0, lambda: progress_var.set(100))
                    root.after(0, lambda: progress_label.config(text="处理完成!"))
                    root.after(0, lambda: messagebox.showinfo("完成", f"处理完成!\n输出文件: {output_file}"))
                    
                except Exception as e:
                    import traceback
                    error_msg = f"处理过程中出错: {e}\n{traceback.format_exc()}"
                    root.after(0, lambda: messagebox.showerror("错误", error_msg))
                finally:
                    # 恢复标准输出
                    sys.stdout = original_stdout
                    # 恢复按钮状态
                    root.after(0, lambda: start_button.config(state=tk.NORMAL))
                    root.after(0, lambda: test_button.config(state=tk.NORMAL))
            
            # 在新线程中运行，避免UI冻结
            thread = threading.Thread(target=run_extraction)
            thread.daemon = True
            thread.start()
        
        def create_test():
            try:
                test_dir = create_test_files()
                folder_var.set(str(test_dir))
                messagebox.showinfo("成功", f"已创建测试文件目录: {test_dir}")
            except Exception as e:
                messagebox.showerror("错误", f"创建测试文件时出错: {e}")
        
        # 添加按钮
        start_button = ttk.Button(button_frame, text="开始处理", command=process_files)
        start_button.pack(side=tk.LEFT, padx=5)
        
        test_button = ttk.Button(button_frame, text="创建测试文件", command=create_test)
        test_button.pack(side=tk.LEFT, padx=5)
        
        quit_button = ttk.Button(button_frame, text="退出", command=root.destroy)
        quit_button.pack(side=tk.RIGHT, padx=5)
        
        # 居中显示窗口
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')
        
        # 启动主循环
        root.mainloop()
        
    except ImportError:
        print("警告: 未安装tkinter库，无法创建GUI界面")
        print("tkinter通常包含在Python标准库中，请检查您的Python安装")
        return False
    except Exception as e:
        print(f"创建GUI界面时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

# 使用示例
if __name__ == "__main__":
    # 如果没有命令行参数，启动GUI
    import sys
    if len(sys.argv) == 1:
        create_gui()
    else:
        # 提取文件名到Excel
        extract_filenames_to_excel("E:/liu/Documents/WPSDrive/201050461/WPS云盘/工作项目/11.电力装备/04上线试用/预算导出/2024预算导出", "文件输出", True)
