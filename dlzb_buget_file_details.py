# -*- coding: utf-8 -*-
"""
批量提取Excel文件中明细表（A8:M8为表头，A列为序号，序号为空为结束），生成明细Excel。
字段：
- 事业部预算编号
- 单据编号
- 序号
- 存货编码
- 存货名称
- 规格型号
- 材质
- 单位
- 预算数量
- 技术标准
- 目标价格类别
- 目标价格
- 行备注
- 源单行号
- 年度合同
- 操作（超链接）
"""

# 后续将逐步实现各功能 

import pandas as pd
import openpyxl
from pathlib import Path
import xlrd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

# 明细表字段
DETAIL_COLUMNS = [
    '事业部预算编号', '单据编号', '序号', '存货编码', '存货名称', '规格型号', '材质', '单位',
    '预算数量', '技术标准', '目标价格类别', '目标价格', '行备注', '源单行号', '年度合同', '操作'
]

def extract_details_from_folder(folder_path, output_file="明细表汇总.xlsx", progress_callback=None, log_callback=None):
    folder = Path(folder_path)
    excel_files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in ['.xls', '.xlsx']]
    if log_callback:
        log_callback(f"共发现{len(excel_files)}个Excel文件待处理。\n")
    all_details = []
    for idx, file in enumerate(excel_files):
        try:
            if file.suffix.lower() == '.xlsx':
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                budget_id = ws.cell(row=4, column=1).value or ''
                doc_id = ws.cell(row=6, column=1).value or ''
                header_row = 8
                row = header_row + 1
                while True:
                    seq = ws.cell(row=row, column=1).value
                    if seq is None or str(seq).strip() == '':
                        break
                    detail = {
                        '事业部预算编号': budget_id,
                        '单据编号': doc_id,
                        '序号': ws.cell(row=row, column=1).value,
                        '存货编码': ws.cell(row=row, column=2).value,
                        '存货名称': ws.cell(row=row, column=3).value,
                        '规格型号': ws.cell(row=row, column=4).value,
                        '材质': ws.cell(row=row, column=5).value,
                        '单位': ws.cell(row=row, column=6).value,
                        '预算数量': ws.cell(row=row, column=7).value,
                        '技术标准': ws.cell(row=row, column=8).value,
                        '目标价格类别': ws.cell(row=row, column=9).value,
                        '目标价格': ws.cell(row=row, column=10).value,
                        '行备注': ws.cell(row=row, column=11).value,
                        '源单行号': ws.cell(row=row, column=12).value,
                        '年度合同': ws.cell(row=row, column=13).value,
                        '操作': str(file.absolute()),
                    }
                    all_details.append(detail)
                    row += 1
            elif file.suffix.lower() == '.xls':
                wb = xlrd.open_workbook(str(file))
                ws = wb.sheet_by_index(0)
                budget_id = ws.cell_value(3, 0) if ws.nrows > 3 else ''
                doc_id = ws.cell_value(5, 0) if ws.nrows > 5 else ''
                header_row = 7
                row = header_row + 1
                while row < ws.nrows:
                    seq = ws.cell_value(row, 0)
                    if seq is None or str(seq).strip() == '':
                        break
                    detail = {
                        '事业部预算编号': budget_id,
                        '单据编号': doc_id,
                        '序号': ws.cell_value(row, 0),
                        '存货编码': ws.cell_value(row, 1),
                        '存货名称': ws.cell_value(row, 2),
                        '规格型号': ws.cell_value(row, 3),
                        '材质': ws.cell_value(row, 4),
                        '单位': ws.cell_value(row, 5),
                        '预算数量': ws.cell_value(row, 6),
                        '技术标准': ws.cell_value(row, 7),
                        '目标价格类别': ws.cell_value(row, 8),
                        '目标价格': ws.cell_value(row, 9),
                        '行备注': ws.cell_value(row, 10),
                        '源单行号': ws.cell_value(row, 11),
                        '年度合同': ws.cell_value(row, 12),
                        '操作': str(file.absolute()),
                    }
                    all_details.append(detail)
                    row += 1
            if log_callback:
                log_callback(f"已处理: {file.name}\n")
            if progress_callback:
                progress_callback((idx + 1) / len(excel_files) * 100)
        except Exception as e:
            if log_callback:
                log_callback(f"处理文件 {file.name} 出错: {e}\n")
    df = pd.DataFrame(all_details, columns=DETAIL_COLUMNS)
    output_path = Path(output_file)
    df.to_excel(output_path, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    op_col = DETAIL_COLUMNS.index('操作') + 1
    for i in range(2, ws.max_row + 1):
        cell = ws.cell(row=i, column=op_col)
        file_path = cell.value
        if file_path:
            cell.value = '打开文件'
            cell.hyperlink = file_path
            cell.style = 'Hyperlink'
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length * 1.2 + 2
    wb.save(output_path)
    if log_callback:
        log_callback(f"明细表已保存到: {output_path.absolute()}\n")
    return output_path

def run_gui():
    root = tk.Tk()
    root.title("明细表批量提取工具")
    root.geometry("700x500")

    style = ttk.Style()
    style.configure("TButton", font=("微软雅黑", 10))
    style.configure("TLabel", font=("微软雅黑", 10))
    style.configure("TEntry", font=("微软雅黑", 10))

    frm = ttk.Frame(root, padding=10)
    frm.pack(fill=tk.BOTH, expand=True)

    # 文件夹选择
    folder_var = tk.StringVar()
    output_var = tk.StringVar(value="明细表汇总.xlsx")

    def browse_folder():
        path = filedialog.askdirectory(title="选择Excel文件夹")
        if path:
            folder_var.set(path)

    ttk.Label(frm, text="Excel文件夹:").grid(row=0, column=0, sticky=tk.W, pady=5)
    folder_entry = ttk.Entry(frm, textvariable=folder_var, width=50)
    folder_entry.grid(row=0, column=1, sticky=tk.W, pady=5)
    ttk.Button(frm, text="浏览...", command=browse_folder).grid(row=0, column=2, padx=5)

    ttk.Label(frm, text="输出文件名:").grid(row=1, column=0, sticky=tk.W, pady=5)
    output_entry = ttk.Entry(frm, textvariable=output_var, width=50)
    output_entry.grid(row=1, column=1, sticky=tk.W, pady=5)

    # 进度条
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(frm, variable=progress_var, maximum=100)
    progress_bar.grid(row=2, column=0, columnspan=3, sticky=tk.EW, pady=10)

    # 日志
    log_text = tk.Text(frm, height=15, font=("Consolas", 9))
    log_text.grid(row=3, column=0, columnspan=3, sticky=tk.NSEW, pady=5)
    frm.rowconfigure(3, weight=1)
    frm.columnconfigure(1, weight=1)

    def log_callback(msg):
        log_text.insert(tk.END, msg)
        log_text.see(tk.END)
        log_text.update_idletasks()

    def progress_callback(val):
        progress_var.set(val)
        root.update_idletasks()

    def start_extract():
        folder = folder_var.get()
        output_file = output_var.get()
        if not folder:
            messagebox.showerror("错误", "请选择Excel文件夹！")
            return
        log_text.delete(1.0, tk.END)
        progress_var.set(0)
        def task():
            try:
                out_path = extract_details_from_folder(folder, output_file, progress_callback, log_callback)
                messagebox.showinfo("完成", f"处理完成！\n输出文件: {out_path}")
            except Exception as e:
                messagebox.showerror("错误", f"处理出错: {e}")
        threading.Thread(target=task, daemon=True).start()

    ttk.Button(frm, text="开始提取", command=start_extract).grid(row=4, column=0, pady=10)
    ttk.Button(frm, text="退出", command=root.destroy).grid(row=4, column=2, pady=10)

    root.mainloop()

if __name__ == "__main__":
    run_gui() 
